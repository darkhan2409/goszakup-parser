"""
Парсер договоров Goszakup с плановыми суммами
Версия 3.0: Два режима экспорта + все улучшения
"""

import requests
from dotenv import load_dotenv
import os
import json
import pandas as pd
from datetime import datetime
import time
import sys
from typing import List, Dict, Optional, Any


load_dotenv()
token = os.getenv('TOKEN')

if not token:
    print("❌ Ошибка: Токен не найден в .env файле!")
    print("Создайте файл .env и добавьте строку: TOKEN=ваш_токен")
    sys.exit(1)

URL = 'https://ows.goszakup.gov.kz/v3/graphql'
headers = {
    'Authorization': f'Bearer {token}',
    'Content-Type': 'application/json'
}

# ========== GraphQL ЗАПРОС ==========
CONTRACTS_QUERY = """
query GetContracts($customerBin: String!, $finYear: Int!, $limit: Int!, $after: Int) {
  Contract(
    filter: {
      customerBin: $customerBin
      finYear: $finYear
    }
    limit: $limit
    after: $after
  ) {
    id
    contractNumberSys
    trdBuyNumberAnno
    trdBuyNameRu
    descriptionRu
    finYear
    contractSum
    signDate

    Supplier {
      nameRu
    }

    RefContractType {
      nameRu
    }

    RefContractStatus {
      nameRu
    }

    FaktTradeMethods {
      nameRu
    }

    ContractUnits {
      plnPointId
      itemPrice
      quantity
      totalSum
    }
  }
}
"""


def fetch_all_contracts(customer_bin: str, fin_year: int, limit: int = 200, max_pages: Optional[int] = None) -> List[Dict[str, Any]]:
    """Получить все договоры с ContractUnits"""
    all_contracts = []
    after = None
    page = 1

    print(f"Получение договоров для БИН {customer_bin}, год {fin_year}...")

    while True:
        if max_pages and page > max_pages:
            print(f"⚠️  Достигнут лимит страниц: {max_pages}")
            break

        variables = {
            'customerBin': customer_bin,
            'finYear': fin_year,
            'limit': limit,
            'after': after
        }

        try:
            response = requests.post(
                URL,
                headers=headers,
                json={'query': CONTRACTS_QUERY, 'variables': variables},
                timeout=30
            )

            if response.status_code != 200:
                print(f"❌ Ошибка API: {response.status_code}")
                if response.status_code == 429:
                    print("⏳ Rate limit, ожидание 60 секунд...")
                    time.sleep(60)
                    continue
                break

            data = response.json()

            if 'errors' in data:
                print(f"❌ GraphQL ошибки: {data['errors']}")
                break

            contracts = data.get('data', {}).get('Contract', [])

            if not contracts:
                break

            all_contracts.extend(contracts)
            print(f"✅ Страница {page}: получено {len(contracts)} записей (всего {len(all_contracts)})")

            after = contracts[-1]['id']
            page += 1
            time.sleep(0.5)

        except requests.exceptions.Timeout:
            print(f"⏱️  Таймаут на странице {page}, повтор через 5 секунд...")
            time.sleep(5)
            continue

        except requests.exceptions.ConnectionError:
            print(f"🔌 Ошибка соединения на странице {page}, повтор через 10 секунд...")
            time.sleep(10)
            continue

        except requests.exceptions.RequestException as e:
            print(f"❌ Ошибка запроса: {e}")
            break

        except Exception as e:
            print(f"❌ Неожиданная ошибка: {e}")
            break

    print(f"\n🎉 Всего получено: {len(all_contracts)} договоров")
    return all_contracts


def fetch_plans_by_ids(plan_ids: List[int]) -> Dict[int, Dict[str, Any]]:
    """Получить планы по списку ID с пагинацией"""
    print(f"\nПолучение планов для {len(plan_ids)} позиций...")

    all_plans = []
    batch_size = 100

    for batch_idx in range(0, len(plan_ids), batch_size):
        batch = plan_ids[batch_idx:batch_idx + batch_size]
        print(f"  Обработка батча {batch_idx // batch_size + 1}/{(len(plan_ids) - 1) // batch_size + 1}...")

        after = None
        page = 1

        while True:
            variables = {
                "ids": batch,
                "limit": 200,
                "after": after
            }

            query_text = """
            query GetPlans($ids: [Int!], $limit: Int!, $after: Int) {
              Plans(
                filter: { id: $ids }
                limit: $limit
                after: $after
              ) {
                id
                nameRu
                count
                price
                amount
                extraDescRu
              }
            }
            """

            try:
                response = requests.post(
                    URL,
                    headers=headers,
                    json={'query': query_text, 'variables': variables},
                    timeout=30
                )

                if response.status_code != 200:
                    if response.status_code == 429:
                        print("⏳ Rate limit, ожидание 60 секунд...")
                        time.sleep(60)
                        continue
                    break

                data = response.json()

                if 'errors' in data:
                    break

                plans = data.get('data', {}).get('Plans', [])

                if not plans:
                    break

                all_plans.extend(plans)

                if len(plans) < 200:
                    break

                after = plans[-1]['id']
                page += 1
                time.sleep(0.3)

            except requests.exceptions.Timeout:
                time.sleep(5)
                continue
            except requests.exceptions.RequestException:
                break

        time.sleep(0.5)

    print(f"✅ Всего планов получено: {len(all_plans)}")
    return {plan['id']: plan for plan in all_plans}


def parse_number(value) -> float:
    """Преобразовать значение в число"""
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        value = value.strip()
        if value:
            try:
                return float(value)
            except ValueError:
                return 0
    return 0


def transform_summary(contracts: List[Dict], plans_dict: Dict) -> pd.DataFrame:
    """
    СВОДНЫЙ РЕЖИМ: 1 договор = 1 строка
    Плановая сумма = сумма всех позиций
    """
    print("\n📊 Трансформация (сводный режим)...")

    rows = []
    stats = {'with_plan': 0, 'without_plan': 0}

    for i, contract in enumerate(contracts, 1):
        supplier_name = contract.get('Supplier', {}).get('nameRu', '') if contract.get('Supplier') else ''
        contract_type = contract.get('RefContractType', {}).get('nameRu', '') if contract.get('RefContractType') else ''
        contract_status = contract.get('RefContractStatus', {}).get('nameRu', '') if contract.get('RefContractStatus') else ''
        procurement_method = contract.get('FaktTradeMethods', {}).get('nameRu', '') if contract.get('FaktTradeMethods') else ''

        description = contract.get('descriptionRu') or contract.get('trdBuyNameRu') or ''

        # Плановая сумма (сумма всех позиций)
        plan_sum = 0
        contract_units = contract.get('ContractUnits', [])

        if contract_units:
            for unit in contract_units:
                pln_point_id = unit.get('plnPointId')
                if pln_point_id and pln_point_id in plans_dict:
                    plan = plans_dict[pln_point_id]
                    plan_sum += parse_number(plan.get('amount'))

        if plan_sum > 0:
            stats['with_plan'] += 1
        else:
            stats['without_plan'] += 1

        # Сумма договора и экономия
        contract_sum = parse_number(contract.get('contractSum'))
        savings = plan_sum - contract_sum if (plan_sum > 0 and contract_sum > 0) else None

        row = {
            '№': i,
            'Номер договора в реестре договоров': contract.get('contractNumberSys', ''),
            'Номер закупки': contract.get('trdBuyNumberAnno', ''),
            'Описание договора': description,
            'Тип договора': contract_type,
            'Статус договора': contract_status,
            'Способ закупки': procurement_method,
            'Финансовый год': contract.get('finYear', ''),
            'Плановая сумма без НДС': plan_sum if plan_sum > 0 else None,
            'Сумма без НДС': contract_sum if contract_sum > 0 else None,
            'Сумма экономии без НДС': savings,
            'Поставщик': supplier_name,
            'Дата заключения': contract.get('signDate', '')
        }

        rows.append(row)

    print(f"\n📊 Статистика:")
    print(f"  С плановой суммой: {stats['with_plan']}")
    print(f"  Без плановой суммы: {stats['without_plan']}")

    df = pd.DataFrame(rows)
    print(f"✅ Преобразовано {len(df)} записей")
    return df


def transform_detailed(contracts: List[Dict], plans_dict: Dict) -> pd.DataFrame:
    """
    ДЕТАЛИЗИРОВАННЫЙ РЕЖИМ: каждая позиция = отдельная строка
    """
    print("\n📊 Трансформация (детализированный режим)...")

    rows = []
    contract_num = 1
    stats = {
        'total_contracts': 0,
        'total_positions': 0,
        'contracts_without_units': 0
    }

    for contract in contracts:
        stats['total_contracts'] += 1

        supplier_name = contract.get('Supplier', {}).get('nameRu', '') if contract.get('Supplier') else ''
        contract_type = contract.get('RefContractType', {}).get('nameRu', '') if contract.get('RefContractType') else ''
        contract_status = contract.get('RefContractStatus', {}).get('nameRu', '') if contract.get('RefContractStatus') else ''
        procurement_method = contract.get('FaktTradeMethods', {}).get('nameRu', '') if contract.get('FaktTradeMethods') else ''
        description = contract.get('descriptionRu') or contract.get('trdBuyNameRu') or ''

        contract_sum = parse_number(contract.get('contractSum'))

        # СТРОКА ЗАГОЛОВОК ДОГОВОРА
        header_row = {
            '_row_type': 'header',
            '№': contract_num,
            'Номер договора': contract.get('contractNumberSys', ''),
            'Номер закупки': contract.get('trdBuyNumberAnno', ''),
            'Описание договора': description,
            'Поставщик': supplier_name,
            'Дата заключения': contract.get('signDate', ''),
            'Тип договора': contract_type,
            'Статус договора': contract_status,
            'Способ закупки': procurement_method,
            'Финансовый год': contract.get('finYear', ''),
            'Общая сумма договора': contract_sum if contract_sum > 0 else None,
            'Наименование позиции': '',
            'Количество': None,
            'Плановая цена за единицу': None,
            'Плановая сумма': None,
            'Сумма по договору': None,
            'Экономия': None
        }
        rows.append(header_row)

        contract_units = contract.get('ContractUnits', [])

        if not contract_units:
            stats['contracts_without_units'] += 1
        else:
            for unit in contract_units:
                stats['total_positions'] += 1

                pln_point_id = unit.get('plnPointId')

                # Данные из Plans
                if pln_point_id and pln_point_id in plans_dict:
                    plan_data = plans_dict[pln_point_id]
                    plan_name = plan_data.get('nameRu', '-')
                    plan_amount = parse_number(plan_data.get('amount'))
                else:
                    plan_name = '-'
                    plan_amount = 0

                unit_sum = parse_number(unit.get('totalSum'))
                unit_price = parse_number(unit.get('itemPrice'))
                unit_qty = parse_number(unit.get('quantity'))

                # Экономия по позиции
                savings = plan_amount - unit_sum if (plan_amount > 0 and unit_sum > 0) else None

                item_row = {
                    '_row_type': 'item',
                    '№': None,
                    'Номер договора': '',
                    'Номер закупки': '',
                    'Описание договора': '',
                    'Поставщик': '',
                    'Дата заключения': '',
                    'Тип договора': '',
                    'Статус договора': '',
                    'Способ закупки': '',
                    'Финансовый год': '',
                    'Общая сумма договора': None,
                    'Наименование позиции': plan_name,
                    'Количество': unit_qty if unit_qty > 0 else None,
                    'Плановая цена за единицу': unit_price if unit_price > 0 else None,
                    'Плановая сумма': plan_amount if plan_amount > 0 else None,
                    'Сумма по договору': unit_sum if unit_sum > 0 else None,
                    'Экономия': savings
                }
                rows.append(item_row)

        contract_num += 1

    print(f"\n📊 Статистика:")
    print(f"  Всего договоров: {stats['total_contracts']}")
    print(f"  Всего позиций: {stats['total_positions']}")
    print(f"  Договоров без позиций: {stats['contracts_without_units']}")
    if stats['total_contracts'] > 0:
        print(f"  Среднее позиций на договор: {stats['total_positions'] / stats['total_contracts']:.1f}")

    df = pd.DataFrame(rows)
    print(f"✅ Преобразовано {len(df)} строк")
    return df


def export_to_excel(df: pd.DataFrame, filename: str, mode: str = 'summary') -> None:
    """Экспорт в Excel с улучшенным форматированием"""
    print(f"\n💾 Экспорт в {filename}...")

    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    # Удалить служебный столбец
    if '_row_type' in df.columns:
        row_types = df['_row_type'].tolist()
        df_export = df.drop(columns=['_row_type'])
    else:
        row_types = []
        df_export = df

    # Числовые колонки для форматирования
    if mode == 'summary':
        numeric_columns = ['Плановая сумма без НДС', 'Сумма без НДС', 'Сумма экономии без НДС']
    else:
        numeric_columns = ['Общая сумма договора', 'Количество', 'Плановая цена за единицу', 
                          'Плановая сумма', 'Сумма по договору', 'Экономия']

    try:
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name='Договоры')

            workbook = writer.book
            worksheet = writer.sheets['Договоры']

            # Границы для ячеек
            thin_border = Border(
                left=Side(style='thin', color='CCCCCC'),
                right=Side(style='thin', color='CCCCCC'),
                top=Side(style='thin', color='CCCCCC'),
                bottom=Side(style='thin', color='CCCCCC')
            )

            # Стиль заголовков
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(name='Times New Roman', bold=True, color='FFFFFF', size=12)

            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = thin_border

            # Найти индексы числовых колонок и колонки экономии
            col_indices = {}
            savings_col_idx = None
            for col_idx, cell in enumerate(worksheet[1], 1):
                if cell.value in numeric_columns:
                    col_indices[cell.value] = col_idx
                if 'экономи' in str(cell.value).lower():
                    savings_col_idx = col_idx

            # Форматирование для сводного режима
            if mode == 'summary':
                # Чередующиеся строки (зебра)
                light_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
                gray_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                
                for row_idx in range(2, worksheet.max_row + 1):
                    # Чередование цвета
                    fill = gray_fill if row_idx % 2 == 0 else light_fill
                    
                    for col_idx in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=row_idx, column=col_idx)
                        cell.border = thin_border
                        cell.fill = fill
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.font = Font(name='Times New Roman', size=12)
                        
                        # Форматирование чисел
                        if col_idx in col_indices.values():
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                        
                        # Цветовое выделение экономии (только отрицательные зна
                        if col_idx == savings_col_idx and cell.value is not None:
                            if isinstance(cell.value, (int, float)):
                                if cell.value > 0:
                                    cell.fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
                                    cell.font = Font(name='Times New Roman', size=12, color='2E7D32', bold=True)
                                elif cell.value < 0:
                                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                                    cell.font = Font(name='Times New Roman', size=12, color='C62828', bold=True)

            # Форматирование для детализированного режима
            elif mode == 'detailed' and row_types:
                header_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                item_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                for idx, row_type in enumerate(row_types, start=2):
                    # Заголовки договоров - серые, позиции - белые
                    fill = header_fill if row_type == 'header' else item_fill
                    
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=idx, column=col)
                        cell.border = thin_border
                        cell.fill = fill
                        cell.alignment = Alignment(vertical='center', wrap_text=True)
                        cell.font = Font(name='Times New Roman', size=12)
                        
                        # Форматирование чисел
                        if col in col_indices.values():
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                        
                        # Цветовое выделение экономии (только отрицательные)
                        if col == savings_col_idx and cell.value is not None:
                            if isinstance(cell.value, (int, float)):
                                if cell.value < 0:
                                    cell.fill = PatternFill(start_color='FFEBEE', end_color='FFEBEE', fill_type='solid')
                                    cell.font = Font(name='Times New Roman', size=12, color='C62828', bold=True)

            # Автоматическая ширина столбцов
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    except (AttributeError, TypeError, ValueError):
                        pass

                adjusted_width = min(max_length + 2, 60)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Закрепить первую строку и включить автофильтр
            worksheet.freeze_panes = 'A2'
            worksheet.auto_filter.ref = worksheet.dimensions

        print(f"✅ Файл сохранён: {filename}")

    except PermissionError:
        print(f"❌ Ошибка: Файл {filename} открыт в другой программе. Закройте файл и повторите попытку.")
        raise
    except Exception as e:
        print(f"❌ Ошибка при сохранении файла: {e}")
        raise


def select_export_mode() -> str:
    """Интерактивный выбор режима экспорта"""
    # Проверить .env
    default_mode = os.getenv('EXPORT_MODE', '').lower()
    if default_mode in ('summary', 'detailed'):
        print(f"📋 Режим из .env: {default_mode}")
        return default_mode

    print("\n" + "=" * 50)
    print("ВЫБОР РЕЖИМА ЭКСПОРТА")
    print("=" * 50)
    print("\n1. 📋 Сводный режим")
    print("   └─ 1 договор = 1 строка")
    print("   └─ Плановая сумма = общая по всем позициям")
    print("   └─ Быстрый обзор всех договоров")

    print("\n2. 📊 Детализированный режим")
    print("   └─ Каждая позиция договора = отдельная строка")
    print("   └─ Видно: наименование, количество, цена")
    print("   └─ Плановая сумма и экономия по каждой позиции")

    while True:
        choice = input("\nВаш выбор (1 или 2): ").strip()

        if choice == '1':
            return 'summary'
        elif choice == '2':
            return 'detailed'
        else:
            print("❌ Неверный выбор. Введите 1 или 2")


def main() -> None:
    """Главная функция"""
    # Параметры из .env
    CUSTOMER_BIN = os.getenv('CUSTOMER_BIN', '020240003361')
    FIN_YEAR = int(os.getenv('FIN_YEAR', '2025'))
    MAX_PAGES = int(os.getenv('MAX_PAGES', '0')) or None
    CACHE_FILE = f'contracts_raw_{FIN_YEAR}.json'

    print("🚀 СТАРТ")
    print(f"📋 Параметры:")
    print(f"   БИН заказчика: {CUSTOMER_BIN}")
    print(f"   Финансовый год: {FIN_YEAR}")
    print(f"   Лимит страниц: {MAX_PAGES if MAX_PAGES else 'без ограничений'}")
    print()

    try:
        # Всегда получаем свежие данные из API
        contracts = fetch_all_contracts(CUSTOMER_BIN, FIN_YEAR, max_pages=MAX_PAGES)
        
        # Сохраняем в кеш для справки
        with open(CACHE_FILE, 'w', encoding='utf-8') as f:
            json.dump(contracts, f, indent=2, ensure_ascii=False)
        print(f"💾 Кеш сохранен: {CACHE_FILE}")

        if not contracts:
            print("⚠️  Договоры не найдены")
            return

        # Собрать все plnPointId
        all_plan_ids = set()
        for contract in contracts:
            for unit in contract.get('ContractUnits', []):
                pln_id = unit.get('plnPointId')
                if pln_id:
                    all_plan_ids.add(pln_id)

        print(f"\n📋 Найдено уникальных plnPointId: {len(all_plan_ids)}")

        # Получить планы
        if all_plan_ids:
            plans_dict = fetch_plans_by_ids(list(all_plan_ids))
        else:
            print("⚠️  plnPointId не найдены")
            plans_dict = {}

        # Выбор режима
        mode = select_export_mode()
        mode_name = 'сводный' if mode == 'summary' else 'детализированный'
        print(f"\n✅ Выбран режим: {mode_name}")

        # Трансформация
        if mode == 'summary':
            df = transform_summary(contracts, plans_dict)
        else:
            df = transform_detailed(contracts, plans_dict)

        # Имя файла
        OUTPUT_FILE = f'contracts_{FIN_YEAR}_{mode}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        # Экспорт
        export_to_excel(df, OUTPUT_FILE, mode=mode)

        print("\n" + "=" * 50)
        print("🎉 ГОТОВО!")
        print("=" * 50)
        print(f"📊 Режим: {mode_name}")
        print(f"📊 Обработано: {len(contracts)} договоров")
        print(f"📊 Строк в таблице: {len(df)}")
        print(f"📄 Файл: {OUTPUT_FILE}")

    except KeyboardInterrupt:
        print("\n\n⚠️  Прервано пользователем")
        sys.exit(0)
    except Exception as e:
        print(f"\n❌ Критическая ошибка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
